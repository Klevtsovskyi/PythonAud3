

import openpyxl
from collections import defaultdict


def create_xlsx(filename, *sheets):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for name, table in sheets:
        ws = wb.create_sheet(name)
        for row in table:
            ws.append(row)

    wb.save(filename)


def create_connections(filename):
    wb = openpyxl.load_workbook(filename)

    proj_ws = wb["projects"]
    dct = defaultdict(list)

    for row in proj_ws.rows:
        dct[row[0].value].append(row[1].value)

    del dct["Project"]
    # print(dct)

    conn_ws = wb.create_sheet("connections")
    row = ("Person 1", "Person 2", "Projects")
    conn_ws.append(row)

    for key, values in dct.items():
        for i in range(len(values)):
            for j in range(i + 1, len(values)):
                row = (values[i], values[j], key)
                conn_ws.append(row)

    wb.save(filename)


if __name__ == '__main__':
    projects = ("projects",
                (("Project", "Person"),
                 (1, "Alex"),
                 (1, "Jack"),
                 (2, "John"),
                 (2, "Alex"),
                 (2, "Stacy"),
                 (3, "Stacy"),
                 (3, "Alex"),
                 (3, "David"),
                 (4, "Alex"),
                 (4, "John"),
                 (5, "Henry")))
    create_xlsx("t23_10.xlsx", projects)
    create_connections("t23_10.xlsx")
